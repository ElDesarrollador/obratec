import shutil
import os
import pandas as pd
from openpyxl import Workbook, load_workbook, cell
from flask import Flask, request, redirect, url_for, send_from_directory, render_template, send_file
from werkzeug.utils import secure_filename
 

CARPETA = os.path.abspath("./XLS/")
EXTENSIONES = set(["xls", "xlsx"])
f_Linea = "./BASE/linea.xlsx"
f_SED = "./BASE/sed.xlsx"


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1] in EXTENSIONES 


app = Flask(__name__)
app.config["CARPETA"] = CARPETA


@app.route("/")
def inicio():
    return render_template("index.html")


@app.route('/LineF')
def formatoL():
	return send_file(f_Linea, as_attachment=True)


@app.route('/SubsF')
def formatoS():
	return send_file(f_SED, as_attachment=True)


@app.route("/line", methods=["GET", "POST"])
def linea():
    if request.method == "POST": 
        f = request.files["archivo_linea"]
        if f and allowed_file(f.filename):
            global Archivo1
            Archivo1 = secure_filename(f.filename)
            f.save(os.path.join(app.config["CARPETA"], Archivo1)) 
            return render_template("up1.html", archivo_linea=Archivo1)
        return render_template("error1.html")
    return render_template("Linea.html")


@app.route("/substation", methods=["GET", "POST"])
def subestacion():
    if request.method == "POST": 
        f = request.files["archivo_sed"]
        if f and allowed_file(f.filename):
            global Archivo2
            Archivo2 = secure_filename(f.filename)
            f.save(os.path.join(app.config["CARPETA"], Archivo2)) 
            return render_template("up2.html", archivo_SED=Archivo2)
        return render_template("error2.html")
    return render_template("SED.html")


@app.route('/grid')
def red():
    return render_template("RED.html", archivo_linea=Archivo1, archivo_SED=Archivo2)


@app.route('/generate', methods=['POST'])
def generar():
    global NombreRED
    NombreRED = request.form.get("nombre")
    global ruta_linea
    global ruta_sed
    ruta_linea=os.path.join(app.config["CARPETA"], Archivo1)
    ruta_sed=os.path.join(app.config["CARPETA"], Archivo2)
    #------------------------------------------------------------ Cargar y limpiar archivo BASE  ------------------------------------------- 
    libro = load_workbook("./BASE/BASE.xlsx")
    hojasvariables=['ElmLne', 'ElmLod', 'ElmTerm','IntGrf','IntGrfcon','ElmSubstat','ElmTr2','StaCubic','RelFuse','StaSwitch']
    for hojavariable in hojasvariables:
        Clear=libro[hojavariable]
        filasClear=Clear.max_row-1
        Clear.delete_rows(2,filasClear)
    #-------------------------------------------------------------- ElmNet ----------------------------------------------------------------  
    ElmNet=libro["ElmNet"]
    ElmNet["B2"]=NombreRED
    #------------------------------------------------------------- IntGrfnet  -------------------------------------------------------------- 
    IntGrfnet=libro["IntGrfnet"]
    IntGrfnet["B2"]=NombreRED
    lines=load_workbook(ruta_linea)
    Hoja1=lines["Hoja1"]
    nlineas=Hoja1.max_row-1
    ElmLne=libro["ElmLne"]
    #-------------------------------------------------------------- ElmLne -----------------------------------------------------------------  
    #1C Elmne
    for i in range(1,nlineas+1):
        ElmLne.cell(row=i+1,column=1,value=730+i)
    #2C Elmne
    for i in range(1,nlineas+1):
        nombre=Hoja1.cell(row=i+1,column=3).value
        ElmLne.cell(row=i+1,column=2,value=nombre)
    #3C Elmne
    for i in range(1,nlineas+1):
        ElmLne.cell(row=i+1,column=3,value=2)
    #4C Elmne
    df1=pd.read_excel("./BASE/BASE.xlsx","TypGeo")
    df2=pd.read_excel("./BASE/BASE.xlsx","TypLne")
    df3=pd.read_excel("./BASE/BASE.xlsx","TypCon")
    #libro.save(filename="./BASE/BASE.xlsx")
    for i in range(1,nlineas+1):
        tension=Hoja1.cell(row=i+1,column=13).value
        if tension==None:
            mater=Hoja1.cell(row=i+1,column=12).value
            idd=df2[df2['loc_name(a:40)']==mater]['ID(a:40)']
            #print(int(idd))
            ElmLne.cell(row=i+1,column=4,value=int(idd)) # No encuentra el tipo subterraneo
        else:
            idd=df1[df1['loc_name(a:40)']==tension]['ID(a:40)'] 
            ElmLne.cell(row=i+1,column=4,value=int(idd)) # Tipo de geometria, revizar si esta en el formato
            #print(int(idd))
        #libro.save(filename="./BASE/BASE.xlsx")
    #5C Elmne
    for i in range(1,nlineas+1):
        longi=float(Hoja1.cell(row=i+1,column=5).value)
        ElmLne.cell(row=i+1,column=5,value=longi/1000)
    #6C Elmne
    for i in range(1,nlineas+1):
        tension=Hoja1.cell(row=i+1,column=13).value
        if tension==None:
            mater=Hoja1.cell(row=i+1,column=12).value
            idd=df2[df2['loc_name(a:40)']==mater]['ID(a:40)']
            ElmLne.cell(row=i+1,column=6,value=int(idd))
            # print(int(idd))
        else:
            cond=Hoja1.cell(row=i+1,column=12).value
            idd=df3[df3['loc_name(a:40)']==cond]['ID(a:40)'] #No encuentra el tipo de conductor
            ElmLne.cell(row=i+1,column=6,value=int(idd))
            # print(int(idd))
        #libro.save(filename="./BASE/BASE.xlsx")
    #7C Elmne
    #8C y 9C Elmne
    for i in range(1,nlineas+1):
        ElmLne.cell(row=i+1,column=8,value=1)
        ElmLne.cell(row=i+1,column=9,value=0)
        refSub=ElmLne.cell(row=i+1,column=1).value
    #----------------------------------------------------------- ElmSubstat -----------------------------------------------------------------
    #1C ElmSubstat
    ElmSubstat=libro["ElmSubstat"]
    seds=load_workbook(ruta_sed)
    Hoja1=seds["Hoja1"]
    nload=Hoja1.max_row-1
    ElmSubstat.cell(row=2,column=1,value=refSub+1)
    for i in range(1,nload+1):
        ElmSubstat.cell(row=i+2,column=1,value=refSub+1+i)
    #2C ElmSubstat
    ElmSubstat.cell(row=2,column=2,value="BARRA0")
    for i in range(1,nload+1):
        nombre=Hoja1.cell(row=i+1,column=3).value
        ElmSubstat.cell(row=i+2,column=2,value=nombre)
    #3C ElmSubstat
    for i in range(1,nload+2):
        ElmSubstat.cell(row=i+1,column=3,value=2)
    #4C ElmSubstat
    Hoja1=lines["Hoja1"]                                       
    for i in range(1,nlineas+1):                              
        cadenatension=Hoja1.cell(row=i+1,column=13).value   
        if cadenatension != None:
            break
    longitudcadena=len(cadenatension)
    if longitudcadena==4:
        kv=cadenatension[0:2]
    else:
        kv=cadenatension[0:4]
    for i in range(1,nload+2):
        ElmSubstat.cell(row=i+1,column=4,value=float(kv))
    #5C ElmSubstat
    lonnodo0=Hoja1.cell(row=2,column=8).value
    ElmSubstat.cell(row=2,column=5,value=lonnodo0)
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        xx=Hoja1.cell(row=i+1,column=6).value
        ElmSubstat.cell(row=i+2,column=5,value=xx)
    #6C ElmSubstat
    Hoja1=lines["Hoja1"]
    latnodo0=Hoja1.cell(row=2,column=9).value
    ElmSubstat.cell(row=2,column=6,value=latnodo0)
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        yy=Hoja1.cell(row=i+1,column=7).value
        ElmSubstat.cell(row=i+2,column=6,value=yy)
    nSubstat=ElmSubstat.max_row-1
    refLod=ElmSubstat.cell(row=nSubstat+1,column=1).value
    #------------------------------------------------------- ElmLod --------------------------------------------------------------------------  
    ElmLod=libro["ElmLod"]
    #1C ElmLod
    for i in range(1,nload+1):
        ElmLod.cell(row=i+1,column=1,value=refLod+i)
    #2C ElmLod
    for i in range(1,nload+1):
        nombre=Hoja1.cell(row=i+1,column=3).value
        ElmLod.cell(row=i+1,column=2,value=nombre)
    #3C ElmLod
    for i in range(1,nload+1):
        idd=ElmSubstat.cell(row=i+2,column=1).value
        ElmLod.cell(row=i+1,column=3,value=idd)
    #4C ElmLod
    for i in range(1,nload+1):
        pot=Hoja1.cell(row=i+1,column=5).value
        #print(pot)
        ElmLod.cell(row=i+1,column=4,value=pot/1000)
    #5C ElmLod
    cosfi=0.95 
    for i in range(1,nload+1):
        ElmLod.cell(row=i+1,column=5,value=float(cosfi))
    #6C ElmLod
    demanda=1 
    sumt=0
    for i in range(1,nload+1):
        sum=Hoja1.cell(row=i+1,column=5).value
        sumt=sumt+sum
    scale=float(demanda)/(sumt/1000)*1
    rscale=round(scale,4)
    for i in range(1,nload+1):
        ElmLod.cell(row=i+1,column=6,value=rscale)
    #7C ElmLod
    for i in range(1,nload+1):
        clien=Hoja1.cell(row=i+1,column=9).value
        ElmLod.cell(row=i+1,column=7,value=clien)
        refTr2=ElmLod.cell(row=i+1,column=1).value
    #----------------------------------------------------------- ElmTr2 --------------------------------------------------------------------
    #1C ElmTr2
    ElmTr2=libro["ElmTr2"]
    for i in range(1,nload+1):
        ElmTr2.cell(row=i+1,column=1,value=i+refTr2)
    #2C ElmTr2
    for i in range(1,nload+1):
        nombre=ElmLod.cell(row=i+1,column=2).value
        concat="TR_"+str(nombre)
        ElmTr2.cell(row=i+1,column=2,value=concat)
    #3C ElmTr2
    for i in range(1,nload+1):
        idd=ElmSubstat.cell(row=i+2,column=1).value
        ElmTr2.cell(row=i+1,column=3,value=idd)
    #4C ElmTr2
    Hoja1=seds["Hoja1"]
    dfTypTr2=pd.read_excel("./BASE/BASE.xlsx","TypTr2")
    for i in range(1,nload+1):
        typ=Hoja1.cell(row=i+1,column=10).value
        # print(typ)
        idd=int(dfTypTr2[dfTypTr2['loc_name(a:40)']==typ]['ID(a:40)']) #Error de no encuentra tipo Tr2
        ElmTr2.cell(row=i+1,column=4,value=idd)
    #5C,6C,7C y 8C ElmTr2
    for i in range(1,nload+1):
        ElmTr2.cell(row=i+1,column=5,value=0)
        ElmTr2.cell(row=i+1,column=6,value=3)
        ElmTr2.cell(row=i+1,column=7,value=0)
        ElmTr2.cell(row=i+1,column=8,value=0)
        refTerm=ElmTr2.cell(row=i+1,column=1).value
    #---------------------------------------------------------- ElmTerm  ----------------------------------------------------------------------- 
    #1C ElmTerm
    ElmTerm=libro["ElmTerm"]
    nnodos=nlineas+1
    for i in range(1,nnodos+2*nload+1):
        ElmTerm.cell(row=i+1,column=1,value=refTerm+i)
    #2C ElmTerm
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        nodo=Hoja1.cell(row=i+1,column=8).value
        ElmTerm.cell(row=3*i-1,column=2,value=str(nodo)+"_1")
        ElmTerm.cell(row=3*i,column=2,value=str(nodo)+"_2")
        ElmTerm.cell(row=3*i+1,column=2,value=nodo)
    Hoja1=lines["Hoja1"]
    ElmTerm.cell(row=3*nload+2,column=2,value=Hoja1.cell(row=2,column=6).value)
    Hoja1=seds["Hoja1"]
    listanodoscarga=[]
    for i in range(1,nload+1):
        nodo=Hoja1.cell(row=i+1,column=8).value
        listanodoscarga.append(nodo)
    Hoja1=lines["Hoja1"]
    j=0
    for i in range(1,nlineas+1):
        nodofin=Hoja1.cell(row=i+1,column=7).value
        if nodofin in listanodoscarga:
            continue
        else:
            j=j+1
            ElmTerm.cell(row=j+3*nload+2,column=2,value=nodofin)
    #3C ElmTerm
    for i in range(1,nload+1):
        idd=ElmSubstat.cell(row=i+2,column=1).value
        ElmTerm.cell(row=3*i-1,column=3,value=idd)
        ElmTerm.cell(row=3*i,column=3,value=idd)
        ElmTerm.cell(row=3*i+1,column=3,value=idd)
    ElmTerm.cell(row=3*nload+2,column=3,value=ElmSubstat.cell(row=2,column=1).value)
    nodosnocarga=nnodos-nload-1
    for i in range(1,nodosnocarga+1):
        ElmTerm.cell(row=i+3*nload+2,column=3,value=2)
    #4C ElmTerm
    for i in range(1,nload+1):
        ElmTerm.cell(row=3*i-1,column=4,value=2)
        ElmTerm.cell(row=3*i,column=4,value=2)
        ElmTerm.cell(row=3*i+1,column=4,value=0)
    ElmTerm.cell(row=3*nload+2,column=4,value=0)
    for i in range(1,nodosnocarga+1):
        ElmTerm.cell(row=i+3*nload+2,column=4,value=1)
    #5C ElmTerm
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        ElmTerm.cell(row=3*i-1,column=5,value=float(kv))
        ph=Hoja1.cell(row=i+1,column=11).value
        if ph==1:
            ElmTerm.cell(row=3*i,column=5,value=0.46)
        else:
            ElmTerm.cell(row=3*i,column=5,value=0.4)
        ElmTerm.cell(row=3*i+1,column=5,value=float(kv))
    ElmTerm.cell(row=3*nload+2,column=5,value=float(kv))
    for i in range(1,nodosnocarga+1):
        ElmTerm.cell(row=i+3*nload+2,column=5,value=float(kv))
    #6C ElmTerm
    for i in range(1,nnodos+2*nload+1):
        ElmTerm.cell(row=i+1,column=6,value=0)
    #7C ElmTerm
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        nfases=Hoja1.cell(row=i+1,column=11).value
        if nfases==3:
            ElmTerm.cell(row=3*i-1,column=7,value=0)
            ElmTerm.cell(row=3*i,column=7,value=0)
            ElmTerm.cell(row=3*i+1,column=7,value=0)
        if nfases==2:
            ElmTerm.cell(row=3*i-1,column=7,value=4)
            ElmTerm.cell(row=3*i,column=7,value=4)
            ElmTerm.cell(row=3*i+1,column=7,value=4)
        if nfases==1:
            ElmTerm.cell(row=3*i-1,column=7,value=6)
            ElmTerm.cell(row=3*i,column=7,value=6)
            ElmTerm.cell(row=3*i+1,column=7,value=6)
    ElmTerm.cell(row=3*nload+2,column=7,value=0)
    j=0
    Hoja1=lines["Hoja1"]
    for i in range(1,nlineas+1):
        nodofin=Hoja1.cell(row=i+1,column=7).value
        if nodofin in listanodoscarga:
            continue
        else:
            j=j+1
            ph=Hoja1.cell(row=i+1,column=15).value
            if ph==1:
                ElmTerm.cell(row=j+3*nload+2,column=7,value=6)
            if ph==2:
                ElmTerm.cell(row=j+3*nload+2,column=7,value=4)
            if ph==3:
                ElmTerm.cell(row=j+3*nload+2,column=7,value=0)
    #8C ElmTerm
    j=0
    Hoja1=lines["Hoja1"]
    for i in range(1,nlineas+1):
        nodofin=Hoja1.cell(row=i+1,column=7).value
        if nodofin in listanodoscarga:
            continue
        else:
            j=j+1
            longps=Hoja1.cell(row=i+1,column=10).value
            ElmTerm.cell(row=j+3*nload+2,column=8,value=longps)
    #9C ElmTerm
    j=0
    for i in range(1,nlineas+1):
        nodofin=Hoja1.cell(row=i+1,column=7).value
        if nodofin in listanodoscarga:
            continue
        else:
            j=j+1
            latgps=Hoja1.cell(row=i+1,column=11).value
            ElmTerm.cell(row=j+3*nload+2,column=9,value=latgps)        
    #10C ElmTerm
    for i in range(1,nnodos+2*nload+1):
        ElmTerm.cell(row=i+1,column=10,value=0)
        refint=ElmTerm.cell(row=i+1,column=1).value
    #------------------------------------------------------ IntGrf -----------------------------------------------------------------------  
    #1C IntGrf
    IntGrf=libro["IntGrf"]
    elementos=nlineas+nload+nodosnocarga+1
    for i in range(1,elementos+1):
        IntGrf.cell(row=i+1,column=1,value=refint+i)
    #2C IntGrf
    for i in range(1,nodosnocarga+2):
        nombre=ElmTerm.cell(row=i+3*nload+1,column=2).value
        con="gno "+str(nombre)
        IntGrf.cell(row=i+1,column=2,value=con)
    for i in range(1,nload+1):
        nombre=ElmLod.cell(row=i+1,column=2).value
        con="gno "+str(nombre)
        IntGrf.cell(row=i+nodosnocarga+2,column=2,value=con)
    for i in range(1,nlineas+1):
        nombre=ElmLne.cell(row=i+1,column=2).value
        con="gno "+str(nombre)
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=2,value=con)
    #3C IntGrf
    for i in range(1,elementos+1):
        IntGrf.cell(row=i+1,column=3,value=3)
    #4C IntGrf
    #ColorSE=imput('ingrese color de SE')
    #ColorTERM=imput('ingrese color de NODOS')
    #ColorLNE=imput('ingrese color de LNE')
    #ColorSED=imput('ingrese color de SED')
    #ColorEEPP=imput('ingrese color de EEPP')
    IntGrf.cell(row=2,column=4,value=13) #Aqui asignar variable ColorSE
    Hoja1=lines["Hoja1"]
    listanodospro=[]
    for i in range(1,nlineas+1):
        fus=Hoja1.cell(row=i+1,column=14).value
        nodopro=Hoja1.cell(row=i+1,column=6).value
        if fus!=0:
            listanodospro.append(nodopro)
    for i in range(1,nodosnocarga+1):
        nodo=ElmTerm.cell(row=i+3*nload+2,column=2).value
        if nodo in listanodospro:
            IntGrf.cell(row=i+2,column=4,value=2) #Aqui asignar variable ColorEEPP
        else:
            IntGrf.cell(row=i+2,column=4,value=1) #Aqui asignar variable ColoTERM
    for i in range(1,nload+1):
        IntGrf.cell(row=i+nodosnocarga+2,column=4,value=13) #Aqui asignar variable ColorSED
    for i in range(1,nlineas+1):
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=4,value=1) #Aqui asignar variable ColorLNE
    #5C y 6C IntGrf
    for i in range(1,elementos+1):
        IntGrf.cell(row=i+1,column=5,value=1)
        IntGrf.cell(row=i+1,column=6,value=1)
    Hoja1=lines["Hoja1"]
    #7C IntGrf
    xmin=0.0001
    for i in range(1,nlineas+1):
        val=float(Hoja1.cell(row=i+1,column=10).value)
        if xmin > val:
            xmin=val
        else:
            xmin=xmin
    escala=32000
    Hoja1=lines["Hoja1"]
    centerx=(Hoja1.cell(row=2,column=8).value - xmin)*escala
    IntGrf.cell(row=2,column=7,value=centerx)
    for i in range(1,nodosnocarga+1):
        lon=ElmTerm.cell(row=i+3*nload+2,column=8).value
        centerx=(lon-xmin)*escala
        IntGrf.cell(row=i+2,column=7,value=centerx)
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        lon=Hoja1.cell(row=i+1,column=6).value
        centerx=(lon-xmin)*escala
        IntGrf.cell(row=i+nodosnocarga+2,column=7,value=centerx)
    Hoja1=lines["Hoja1"]
    for i in range(1,nlineas+1):
        prom=((Hoja1.cell(row=i+1,column=8).value + Hoja1.cell(row=i+1,column=10).value)/2-xmin)*escala
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=7,value=prom)
    #8C IntGrf
    Hoja1=lines["Hoja1"]
    ymin=0.0001
    for i in range(1,nnodos):
        val=float(Hoja1.cell(row=i+1,column=11).value)
        if ymin > val:
            ymin=val
        else:
            ymin=ymin
    Hoja1=lines["Hoja1"]
    centery=(Hoja1.cell(row=2,column=9).value - ymin)*escala
    IntGrf.cell(row=2,column=8,value=centery)
    for i in range(1,nodosnocarga+1):
        lat=ElmTerm.cell(row=i+3*nload+2,column=9).value
        centery=(lat-ymin)*escala
        IntGrf.cell(row=i+2,column=8,value=centery)
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        lat=Hoja1.cell(row=i+1,column=7).value
        centery=(lat-ymin)*escala
        IntGrf.cell(row=i+nodosnocarga+2,column=8,value=centery)
    Hoja1=lines["Hoja1"]
    for i in range(1,nlineas+1):
        prom=((Hoja1.cell(row=i+1,column=9).value + Hoja1.cell(row=i+1,column=11).value)/2-ymin)*escala
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=8,value=prom)
    #9C IntGrf
    IntGrf.cell(row=2,column=9,value="PrimSubstat")
    #TipoSED=imput('ingrese tipo grafico de SED (TRANGULAR O CIRCULAR)')
    #TipoEEPP=imput('ingrese tipo grafico de EEPP(CIRCULAR,CIRCULARFONDO, CIRCULAR-R)')
    for i in range(1,nodosnocarga+1):
        nodo=ElmTerm.cell(row=i+3*nload+2,column=2).value
        if nodo in listanodospro:
            IntGrf.cell(row=i+2,column=9,value="CircTerm1") #Asignar variable de tipo grafico EEPP
        else:
            IntGrf.cell(row=i+2,column=9,value="PointTerm")
    for i in range(1,nload+1):
        IntGrf.cell(row=i+nodosnocarga+2,column=9,value="SecSubProd") #Asignar variable de tipo grafico SED
    for i in range(1,nlineas+1):
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=9,value="d_lin") 
    #10C IntGrf
    IntGrf.cell(row=2,column=10,value=ElmSubstat.cell(row=2,column=1).value)
    for i in range(1,nodosnocarga+1):
        DO=ElmTerm.cell(row=i+3*nload+2,column=1).value
        IntGrf.cell(row=i+2,column=10,value=DO)
    for i in range(1,nload+1):
        DO=ElmSubstat.cell(row=i+2,column=1).value
        IntGrf.cell(row=i+nodosnocarga+2,column=10,value=DO)
    for i in range(1,nlineas+1):
        DO=ElmLne.cell(row=i+1,column=1).value
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=10,value=DO)
    #11C 
    for i in range(1,elementos+1):
        IntGrf.cell(row=i+1,column=11,value=0)
        refint=IntGrf.cell(row=i+1,column=1).value
    #12C y 13C IntGrf
    #TamañoSE=imput('ingrese tamaño grafico de SE')
    #TamañoSED=imput('ingrese tamaño grafico de SED')
    #TamañoEEPP=imput('ingrese tamaño grafico de EEPP')
    IntGrf.cell(row=2,column=12,value=1) #Asignar variable TamañoSE
    IntGrf.cell(row=2,column=13,value=1)
    for i in range(1,nodosnocarga+1):
        nodo=ElmTerm.cell(row=i+3*nload+2,column=2).value
        if nodo in listanodospro:
            IntGrf.cell(row=i+2,column=12,value=1) #Asignar variable tamañoEEPP
            IntGrf.cell(row=i+2,column=13,value=1)
        else:
            IntGrf.cell(row=i+2,column=12,value=1)
            IntGrf.cell(row=i+2,column=13,value=1) 
    for i in range(1,nload+1):
        IntGrf.cell(row=i+nodosnocarga+2,column=12,value=2) #Asignar variable tamañoSED
        IntGrf.cell(row=i+nodosnocarga+2,column=13,value=2)
    for i in range(1,nlineas+1):
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=12,value=1)
        IntGrf.cell(row=i+nodosnocarga+nload+2,column=13,value=1)

    #------------------------------------------------------- IntGrfCon ------------------------------------------------------------------ 
    #1C IntGrfcon
    ncon=nlineas*2
    IntGrfcon=libro["IntGrfcon"]
    for i in range(1,ncon+1):
        IntGrfcon.cell(row=i+1,column=1,value=refint+i)
    #2C IntGrfcon
    for i in range(1,nlineas+1):
        IntGrfcon.cell(row=2*i,column=2,value="GCO_1")
        IntGrfcon.cell(row=2*i+1,column=2,value="GCO_2")
    #3C IntGrfcon
    for i in range(1,nlineas+1):
        id=IntGrf.cell(row=i+nodosnocarga+nload+2,column=1).value
        IntGrfcon.cell(row=2*i,column=3,value=id)
        IntGrfcon.cell(row=2*i+1,column=3,value=id)
    #4C IntGrfcon
    for i in range(1,2*nlineas+1):
        IntGrfcon.cell(row=i+1,column=4,value=2)
    #5C IntGrfcon
    for i in range(1,nlineas+1):
        rx0=IntGrf.cell(row=i+nodosnocarga+nload+2,column=7).value
        IntGrfcon.cell(row=2*i,column=5,value=rx0)
        IntGrfcon.cell(row=2*i+1,column=5,value=rx0)
    #6C IntGrfcon
    Hoja1=lines["Hoja1"]
    for i in range(1,nlineas+1):
        gc1=Hoja1.cell(row=i+1,column=8).value
        gc2=Hoja1.cell(row=i+1,column=10).value
        IntGrfcon.cell(row=2*i,column=6,value=(gc1-xmin)*escala)
        IntGrfcon.cell(row=2*i+1,column=6,value=(gc2-xmin)*escala)
    #7C IntGrfcon
    for i in range(1,2*nlineas+1):
        IntGrfcon.cell(row=i+1,column=7,value=2)
    #8C IntGrfcon
    for i in range(1,nlineas+1):
        ry0=IntGrf.cell(row=i+nodosnocarga+nload+2,column=8).value
        IntGrfcon.cell(row=2*i,column=8,value=ry0)
        IntGrfcon.cell(row=2*i+1,column=8,value=ry0)
    #9C IntGrfcon
    Hoja1=lines["Hoja1"]
    for i in range(1,nlineas+1):
        gc1=Hoja1.cell(row=i+1,column=9).value
        gc2=Hoja1.cell(row=i+1,column=11).value
        IntGrfcon.cell(row=2*i,column=9,value=(gc1-ymin)*escala)
        IntGrfcon.cell(row=2*i+1,column=9,value=(gc2-ymin)*escala)
        reffuse=IntGrfcon.cell(row=2*i+1,column=1).value
    libro.save(filename="./BASE/BASE.xlsx")
    #--------------------------------------------------------RelFuse-------------------------------------------------------------------- 
    #1C RelFuse
    RelFuse=libro["RelFuse"]
    Hoja1=lines["Hoja1"]
    nfuse=0
    for i in range(1,nlineas+1):
        conf=Hoja1.cell(row=i+1,column=14).value
        if conf!=0:
            nfuse=nfuse+1
    for i in range(1,nfuse+nload+1):
        RelFuse.cell(row=i+1,column=1,value=i+reffuse)
    #2C RelFuse
    for i in range(1,nload+1):
        nombre=ElmLod.cell(row=i+1,column=2).value
        con="Fus_"+str(nombre)
        RelFuse.cell(row=i+1,column=2,value=con)
    j=0
    for i in range(1,nlineas+1):
        fusename=Hoja1.cell(row=i+1,column=14).value
        if fusename!=0:
            j=j+1
            RelFuse.cell(row=j+nload+1,column=2,value=fusename)
    #3C RelFuse (First Part)
    for i in range(1,nload+1):
        idd=ElmSubstat.cell(row=i+2,column=1).value
        RelFuse.cell(row=i+1,column=3,value=idd)
    #4C RelFuse
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        pot=Hoja1.cell(row=i+1,column=5).value
        irat=1.3*pot/(1.7321*float(kv))
        if irat<1:
            RelFuse.cell(row=i+1,column=4,value=730)
            continue
        if irat<2 and irat>1:
            RelFuse.cell(row=i+1,column=4,value=729)
            continue
        if irat<3 and irat>2:
            RelFuse.cell(row=i+1,column=4,value=728)
            continue
        if irat<5 and irat>3:
            RelFuse.cell(row=i+1,column=4,value=727)
            continue
        if irat<6 and irat>5:
            RelFuse.cell(row=i+1,column=4,value=726)
            continue
        if irat<7 and irat>6:
            RelFuse.cell(row=i+1,column=4,value=725)
            continue
        if irat<8 and irat>7:
            RelFuse.cell(row=i+1,column=4,value=724)
            continue
        if irat<10 and irat>8:
            RelFuse.cell(row=i+1,column=4,value=723)
            continue
        if irat<12 and irat>10:
            RelFuse.cell(row=i+1,column=4,value=722)
            continue
        if irat<15 and irat>12:
            RelFuse.cell(row=i+1,column=4,value=721)
            continue
        if irat<20 and irat>15:
            RelFuse.cell(row=i+1,column=4,value=720)
            continue
        if irat<30 and irat>20:
            RelFuse.cell(row=i+1,column=4,value=719)
            continue
        if irat<40 and irat>30:
            RelFuse.cell(row=i+1,column=4,value=718)
            continue
        if irat<50 and irat>40:
            RelFuse.cell(row=i+1,column=4,value=717)
            continue
        if irat<65 and irat>50:
            RelFuse.cell(row=i+1,column=4,value=716)
            continue
        if irat<80 and irat>65:
            RelFuse.cell(row=i+1,column=4,value=715)
            continue
        if irat<100 and irat>80:
            RelFuse.cell(row=i+1,column=4,value=714)
            continue
        if irat<140 and irat>100:
            RelFuse.cell(row=i+1,column=4,value=713)
            continue
        if irat<200 and irat>140:
            RelFuse.cell(row=i+1,column=4,value=712)
            continue
    for i in range(1,nfuse+1):
        RelFuse.cell(row=i+nload+1,column=4,value=712)
    #5C,6C,7C y 8C RelFuse
    for i in range(1,nfuse+nload+1):
        RelFuse.cell(row=i+1,column=5,value=1)
        RelFuse.cell(row=i+1,column=6,value="fus")
        RelFuse.cell(row=i+1,column=7,value=0)
        #RelFuse.cell(row=i+1,column=8,value=3)# AQUI MEJORAR EL CODIGO,YA K A TODOO LOS FUSIBLES LE PONE TRIFASICO
        refSta=RelFuse.cell(row=i+1,column=1).value
    #************************************
    Hoja1=seds["Hoja1"]
    for i in range(1,nload+1):
        ph=Hoja1.cell(row=i+1,column=11).value
        RelFuse.cell(row=i+1,column=8,value=int(ph))
    Hoja1=lines["Hoja1"]
    j=0
    for i in range(1,nlineas+1):
        fusename=Hoja1.cell(row=i+1,column=14).value
        ph=Hoja1.cell(row=i+1,column=15).value
        if fusename!=0:
            j=j+1
            RelFuse.cell(row=j+nload+1,column=8,value=ph)
    #**********************************
    #---------------------------------------------------------- StaCubic ----------------------------------------------------------------- 
    #1C StaCubic
    StaCubic=libro["StaCubic"]
    for i in range(1,2*nlineas+5*nload+1):
        StaCubic.cell(row=i+1,column=1,value=refSta+i)
    #2C StaCubic
    for i in range(1,nlineas+1):
        id=ElmLne.cell(row=i+1,column=1).value
        StaCubic.cell(row=2*i,column=2,value="Cub_"+str(id))
        StaCubic.cell(row=2*i+1,column=2,value="Cub_"+str(id))
    for i in range(1,nload+1):
        id=RelFuse.cell(row=i+1,column=1).value
        StaCubic.cell(row=2*i+2*nlineas,column=2,value="Cub_"+str(id))
        StaCubic.cell(row=2*i+1+2*nlineas,column=2,value="Cub_"+str(id))
    for i in range(1,nload+1):
        id=ElmTr2.cell(row=i+1,column=1).value
        StaCubic.cell(row=2*i+2*nlineas+2*nload,column=2,value="Cub_"+str(id))
        StaCubic.cell(row=2*i+1+2*nlineas+2*nload,column=2,value="Cub_"+str(id))
    for i in range(1,nload+1):
        id=ElmLod.cell(row=i+1,column=1).value
        StaCubic.cell(row=i+2*nlineas+4*nload+1,column=2,value="Cub_"+str(id))
    #3C StaCubic
    libro.save(filename="./BASE/BASE.xlsx")
    Hoja1=lines["Hoja1"]
    dfterm=pd.read_excel("./BASE/BASE.xlsx","ElmTerm")
    for i in range(1,nlineas+1):
        nodofin=Hoja1.cell(row=i+1,column=7).value
        nodoini=Hoja1.cell(row=i+1,column=6).value
        idf=int(dfterm[dfterm['loc_name(a:40)']==nodofin]['ID(a:40)'])
        idi=int(dfterm[dfterm['loc_name(a:40)']==nodoini]['ID(a:40)']) #PROBLEMAS DE NOMBRES COINCIDENTES
        StaCubic.cell(row=2*i,column=3,value=idf)
        StaCubic.cell(row=2*i+1,column=3,value=idi)
        # print(nodofin)
        #print(type(nodofin))
        # print(idf)
        #print(nodoini)
        # print(type(idf))
        #libro.save(filename="./BASE/BASE.xlsx")
    for i in range(1,nload+1):
        id1=ElmTerm.cell(row=3*i-1,column=1).value
        id2=ElmTerm.cell(row=3*i+1,column=1).value
        StaCubic.cell(row=2*i+2*nlineas,column=3,value=id1)
        StaCubic.cell(row=2*i+1+2*nlineas,column=3,value=id2)
    for i in range(1,nload+1):
        id1=ElmTerm.cell(row=3*i-1,column=1).value
        id2=ElmTerm.cell(row=3*i,column=1).value
        StaCubic.cell(row=2*i+2*nlineas+2*nload,column=3,value=id2)
        StaCubic.cell(row=2*i+1+2*nlineas+2*nload,column=3,value=id1)
    for i in range(1,nload+1):
        idd=ElmTerm.cell(row=3*i,column=1).value
        StaCubic.cell(row=i+2*nlineas+4*nload+1,column=3,value=idd)

    #4C StaCubic
    for i in range(1,nlineas+2*nload+1):
        StaCubic.cell(row=2*i,column=4,value=1)
        StaCubic.cell(row=2*i+1,column=4,value=0)
    for i in range(1,nload+1):
        StaCubic.cell(row=i+2*nlineas+4*nload+1,column=4,value=0)
    #5C StaCubic
    for i in range(1,nlineas+1):
        id=ElmLne.cell(row=i+1,column=1).value
        StaCubic.cell(row=2*i,column=5,value=int(id))
        StaCubic.cell(row=2*i+1,column=5,value=int(id))
    for i in range(1,nload+1):
        id=RelFuse.cell(row=i+1,column=1).value
        StaCubic.cell(row=2*i+2*nlineas,column=5,value=int(id))
        StaCubic.cell(row=2*i+1+2*nlineas,column=5,value=int(id))
    for i in range(1,nload+1):
        id=ElmTr2.cell(row=i+1,column=1).value
        StaCubic.cell(row=2*i+2*nlineas+2*nload,column=5,value=int(id))
        StaCubic.cell(row=2*i+1+2*nlineas+2*nload,column=5,value=int(id))
    for i in range(1,nload+1):
        id=ElmLod.cell(row=i+1,column=1).value
        StaCubic.cell(row=i+2*nlineas+4*nload+1,column=5,value=int(id))
        refswi=StaCubic.cell(row=i+2*nlineas+4*nload+1,column=1).value
    #Last row StaCubic
    StaCubic.cell(row=2*nlineas+5*nload+2,column=1,value=refswi+1)
    StaCubic.cell(row=2*nlineas+5*nload+2,column=2,value="CubicXnet")
    StaCubic.cell(row=2*nlineas+5*nload+2,column=3,value=ElmTerm.cell(row=3*nload+2,column=1).value)
    StaCubic.cell(row=2*nlineas+5*nload+2,column=4,value=0)
    StaCubic.cell(row=2*nlineas+5*nload+2,column=5,value=1000000000)
    libro.save(filename="./BASE/BASE.xlsx")
    #3C RelFuse (Second Part)
    dfLne=pd.read_excel("./BASE/BASE.xlsx","ElmLne")
    dfcubic=pd.read_excel("./BASE/BASE.xlsx","StaCubic")
    j=0
    for i in range(1,nlineas+1):
        fuseActive=Hoja1.cell(row=i+1,column=14).value
        if fuseActive!=0:
            pi=Hoja1.cell(row=i+1,column=3).value
            idl=int(dfLne[dfLne['loc_name(a:40)']==pi]['ID(a:40)'])
            idc=int(dfcubic[(dfcubic['obj_id(p)']==idl) & (dfcubic['obj_bus(i)']==0)]['ID(a:40)'])
            j=j+1
            RelFuse.cell(row=j+1+nload,column=3,value=idc)
    #--------------------------------------------------StaSwitch---------------------------------------------------------------------- 
    #1C StaSwitch
    StaSwitch=libro["StaSwitch"]
    for i in range(1,nfuse+1):
        StaSwitch.cell(row=i+1,column=1,value=i+refswi+1)
    #2C StaSwitch
    for i in range(1,nfuse+1):
        StaSwitch.cell(row=i+1,column=2,value="Switch")
    #3C StaSwitch
    for i in range(1,nfuse+1):
        id=RelFuse.cell(row=i+1+nload,column=3).value
        StaSwitch.cell(row=i+1,column=3,value=id)
    #4C y 5C StaSwitch
    for i in range(1,nfuse+1):
        StaSwitch.cell(row=i+1,column=4,value=1)
        StaSwitch.cell(row=i+1,column=5,value=1)
    #----------------------------------------------------------- ElmXnet ----------------------------------------
    ElmXnet=libro["ElmXnet"]
    ElmXnet.cell(row=2,column=3,value=ElmSubstat.cell(row=2,column=1).value)
    libro.save(filename="./BASE/BASE.xlsx")
    #-------------------------------------------------------------- Crea el archivo .dgs ------------------------------------
    dgs=open('./BASE/DGS.dgs','w')
    hojas=libro.sheetnames
    for hoja in hojas:
        shett=libro[hoja]
        nf=shett.max_row+1
        nc=shett.max_column+1
        dgs.write('$$')
        dgs.write(hoja)
        dgs.write(";")
        for i in range(1,nf):
            for j in range(1,nc):
                value=str(shett.cell(row=i,column=j).value)
                if value=="None":
                    value=""
                    dgs.write(value)
                else:
                    dgs.write(value)
                if j==nc-1:
                    break
                dgs.write(";")
            dgs.write('\n')
        dgs.write('\n')
    dgs.close()
    #------------------------------------------------------------- Crea el archivo .dgs en el escritorio---------------------------------
    fuente="./BASE/DGS.dgs"
    destino="./CONV/DGS.dgs"                    
    shutil.copy(fuente, destino)
    global arch_DGS
    arch_DGS="./CONV/"+NombreRED+".dgs"
    os.rename("./CONV/DGS.dgs",arch_DGS)
    #-------------------------------------------------------------- Deja contenido BASE limpio --------------------------------------------
    libro=load_workbook("./BASE/BASE.xlsx")
    hojasvariables=['ElmLne', 'ElmLod', 'ElmTerm','IntGrf','IntGrfcon','ElmSubstat','ElmTr2','StaCubic','RelFuse','StaSwitch']
    for hojavariable in hojasvariables:
        Clear=libro[hojavariable]
        filasClear=Clear.max_row-1
        Clear.delete_rows(2,filasClear)
    libro.save(filename="./BASE/BASE.xlsx")

    #-----INICIA FUNCIÓN PARA GENERAR KML-----
    #---------------------------------------------------- Carga de archivos Lineas y Seds ---------------------------------------------
    lines=load_workbook(ruta_linea)
    Hoja1=lines["Hoja1"]
    nlineas=Hoja1.max_row-1
    seds=load_workbook(ruta_sed)
    Hoja2=seds["Hoja1"]
    nload=Hoja2.max_row-1
    #---------------------------------------------------- Creacion Archivo TXT -------------------------------------
    KML = open('KML.kml', 'w')
    KML.write('<?xml version="1.0" encoding="UTF-8"?> \n')
    KML.write('<kml xmlns="http://www.opengis.net/kml/2.2" xmlns:gx="http://www.google.com/kml/ext/2.2" xmlns:kml="http://www.opengis.net/kml/2.2" xmlns:atom="http://www.w3.org/2005/Atom"> \n')
    KML.write('<Document><name> \n')
    KML.write(NombreRED)
    KML.write(' \n')
    KML.write('</name><Style id="s_ylw-pushpin"> \n')
    KML.write('<LineStyle><width>3</width></LineStyle></Style><Style id="sn_triangle"><IconStyle><scale>1.2</scale><Icon> \n')
    KML.write('<href>http://maps.google.com/mapfiles/kml/shapes/triangle.png</href> \n')
    KML.write('</Icon></IconStyle></Style><StyleMap id="msn_triangle"><Pair><key>normal</key><styleUrl>#sn_triangle</styleUrl></Pair><Pair><key>highlight</key> \n')
    KML.write('<styleUrl>#sh_triangle</styleUrl></Pair></StyleMap><StyleMap id="m_ylw-pushpin0"><Pair><key>normal</key><styleUrl>#s_ylw-pushpin0</styleUrl> \n')
    KML.write('</Pair><Pair><key>highlight</key><styleUrl>#s_ylw-pushpin_hl2</styleUrl></Pair></StyleMap><Style id="s_ylw-pushpin0"><IconStyle><scale>1.2</scale><Icon> \n')
    KML.write('<href>http://maps.google.com/mapfiles/kml/shapes/homegardenbusiness.png</href> \n')
    KML.write('</Icon><hotSpot x="0.5" y="0" xunits="fraction" yunits="fraction"/></IconStyle></Style><Style id="s_ylw-pushpin_hl2"><IconStyle><scale>1.4</scale><Icon> \n')
    KML.write('<href>http://maps.google.com/mapfiles/kml/shapes/homegardenbusiness.png</href> \n')
    KML.write('</Icon><hotSpot x="0.5" y="0" xunits="fraction" yunits="fraction"/></IconStyle></Style> \n')	
    KML.write('<Style id="s_ylw-pushpin_hl"><IconStyle><scale>1.3</scale><Icon><href>http://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png</href></Icon> \n')
    KML.write('<hotSpot x="20" y="2" xunits="pixels" yunits="pixels"/></IconStyle><LineStyle><width>3</width></LineStyle></Style> \n')
    KML.write('<Style id="sh_triangle"><IconStyle><scale>1.4</scale><Icon><href>http://maps.google.com/mapfiles/kml/shapes/triangle.png</href></Icon></IconStyle></Style> \n')          
    KML.write('<StyleMap id="m_ylw-pushpin"><Pair><key>normal</key><styleUrl>#s_ylw-pushpin</styleUrl></Pair><Pair><key>highlight</key><styleUrl>#s_ylw-pushpin_hl</styleUrl></Pair></StyleMap> \n')
    KML.write('<Folder><name> \n')
    KML.write(NombreRED)
    KML.write(' \n')
    KML.write('</name><open>1</open> \n')
    #---------------------------------------------------- Escritura de Lineas en KML -------------------------------------
    for i in range(1,nlineas+1):
        KML.write('<Placemark><name> \n')
        vano_id = Hoja1.cell(row=i+1,column=3).value
        KML.write(str(vano_id))
        KML.write(' \n')
        KML.write('</name><styleUrl>#m_ylw-pushpin</styleUrl><LineString><tessellate>1</tessellate><coordinates> \n')
        coorX1 = str(Hoja1.cell(row=i+1,column=8).value)
        coorY1 = str(Hoja1.cell(row=i+1,column=9).value)
        coorX2 = str(Hoja1.cell(row=i+1,column=10).value)
        coorY2 = str(Hoja1.cell(row=i+1,column=11).value)
        coordenadas = coorX1+','+coorY1+',0,'+coorX2+','+coorY2+',0'
        KML.write(coordenadas)
        KML.write(' \n')
        KML.write('</coordinates></LineString></Placemark> \n')
    #---------------------------------------------------- Escritura de Seds en KML -------------------------------------
    for i in range(1,nload+1):
        KML.write('<Placemark><name> \n')
        sed_id = Hoja2.cell(row=i+1,column=3).value
        KML.write(str(sed_id))
        KML.write(' \n')
        KML.write('</name><styleUrl>#msn_triangle</styleUrl><Point><gx:drawOrder>1</gx:drawOrder><coordinates> \n')
        coorX1 = str(Hoja2.cell(row=i+1,column=6).value)
        coorY1 = str(Hoja2.cell(row=i+1,column=7).value)
        coordenadas = coorX1+','+coorY1+',0'
        KML.write(coordenadas)
        KML.write(' \n')
        KML.write('</coordinates></Point></Placemark> \n')
    KML.write('</Folder></Document></kml>')
    KML.close()
    #---------------------------------------------------- Crea el archivo .kml en el escritorio---------------------------------
    fuente="./KML.kml"
    destino="./CONV/KML.kml"                    
    shutil.copy(fuente, destino)
    global arch_KML
    arch_KML="./CONV/"+NombreRED+".kml"
    os.rename("./CONV/KML.kml",arch_KML)

    return render_template('dwnld.html')


@app.route('/dnld1')
def descargarDGS():
	return send_file(arch_DGS, as_attachment=True)


@app.route('/dnld2')
def descargarKML():
	return send_file(arch_KML, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True) 